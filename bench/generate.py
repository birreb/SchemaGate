"""Generate the benchmark documents and the rows a correct ingestion produces.

Everything is synthetic and seeded, so the set can be regenerated identically
and published without anyone's real invoice in it. The documents are built to
carry the things that go wrong when a document is mapped onto a table: a buyer
VAT number printed next to the seller's, three dates of which one is wanted,
discount and shipping lines between the line items and the total, European
number formats, dates written as words, and a share of invoices whose printed
total does not add up.

Eight visual families for invoices and four for statements, so that no two
suppliers' documents share a layout: a typewriter ERP printout, a German formal
letter with an address window, a US letter with Bill To and Ship To, a Swedish
invoice with a payment slip and an OCR number, and so on. Fonts come from the
system where available and fall back to the PDF core fonts, so the text layer
is the same on every machine even if the glyphs differ.

    uv run python bench/generate.py

Writes `bench/data/` and `bench/data/manifest.jsonl`. One manifest line per
document: the file, the target table, the rows a correct reading produces, and
how many values the validation gate is expected to reject.
"""

from __future__ import annotations

import csv
import datetime as dt
import hashlib
import io
import json
import logging
import random
from dataclasses import asdict, dataclass, field
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any

from fpdf import FPDF
from openpyxl import Workbook
from PIL import Image, ImageDraw, ImageFilter, ImageFont

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data"
SEED = 20260902

# fontTools reports every table it cannot subset. Nothing to act on.
logging.getLogger("fontTools").setLevel(logging.ERROR)

CENT = Decimal("0.01")

BUYER = {
    "name": "Halvard Industri AB",
    "vat": "SE559012345601",
    "address": ["Verkstadsgatan 14", "721 30 Västerås", "Sweden"],
    "ship": ["Halvard Industri AB, Godsmottagning", "Lagervägen 3", "721 38 Västerås", "Sweden"],
}

# One entry per fictional supplier. `vat` is what a correct reading of the
# seller VAT number produces: None for the US supplier, whose invoice carries
# an EIN, which is not a VAT number and which the column comment excludes.
SUPPLIERS = [
    {
        "name": "Northgate Supply Co.",
        "locale": "sv",
        "vat": "SE556000000001",
        "ccy": "SEK",
        "address": ["Industrivägen 8", "633 42 Eskilstuna"],
        "terms": 30,
        "bank": "Bankgiro 5432-1098",
        "reg": "556000-0000",
        "customer_no": "K-10442",
        "accent": (31, 78, 121),
    },
    {
        "name": "Björkman & Söner AB",
        "locale": "sv",
        "vat": "SE556234567801",
        "ccy": "SEK",
        "address": ["Hamngatan 2", "411 06 Göteborg"],
        "terms": 30,
        "bank": "Bankgiro 320-9917",
        "reg": "556234-5678",
        "customer_no": "40871",
        "accent": (140, 60, 30),
    },
    {
        "name": "Lindqvist Elektronik AB",
        "locale": "sv",
        "vat": "SE556876543201",
        "ccy": "SEK",
        "address": ["Kvarnvägen 19", "587 31 Linköping"],
        "terms": 14,
        "bank": "Plusgiro 48 22 11-3",
        "reg": "556876-5432",
        "customer_no": "C00213",
        "accent": (0, 110, 100),
    },
    {
        "name": "Rheinwerk Maschinenbau GmbH",
        "locale": "de",
        "vat": "DE811234567",
        "ccy": "EUR",
        "address": ["Industriestraße 44", "50735 Köln"],
        "terms": 30,
        "bank": "IBAN DE89 3704 0044 0532 0130 00",
        "reg": "HRB 12345 Amtsgericht Köln",
        "customer_no": "20077",
        "accent": (60, 60, 60),
    },
    {
        "name": "Hofmann Verpackungen GmbH",
        "locale": "de",
        "vat": "DE129876543",
        "ccy": "EUR",
        "address": ["Am Hafen 3", "20457 Hamburg"],
        "terms": 14,
        "bank": "IBAN DE02 2005 0550 1234 5678 90",
        "reg": "HRB 98765 Amtsgericht Hamburg",
        "customer_no": "K-3391",
        "accent": (170, 100, 20),
    },
    {
        "name": "Kestrel Components Ltd",
        "locale": "gb",
        "vat": "GB123456789",
        "ccy": "GBP",
        "address": ["Unit 4, Riverside Park", "Sheffield S9 2AB"],
        "terms": 30,
        "bank": "Sort code 12-34-56, account 87654321",
        "reg": "Company no. 04512345",
        "customer_no": "HALV01",
        "accent": (90, 40, 110),
    },
    {
        "name": "Albion Fasteners Ltd",
        "locale": "gb",
        "vat": "GB987654321",
        "ccy": "GBP",
        "address": ["12 Mill Lane", "Leeds LS1 4DT"],
        "terms": 30,
        "bank": "Sort code 40-11-22, account 11223344",
        "reg": "Company no. 07788990",
        "customer_no": "A-2210",
        "accent": (0, 90, 60),
    },
    {
        "name": "Great Lakes Tooling Inc.",
        "locale": "us",
        "vat": None,
        "ccy": "USD",
        "address": ["2200 Commerce Dr", "Grand Rapids, MI 49503"],
        "terms": 30,
        "ein": "38-2947103",
        "bank": "Wire: Fifth Third Bank, acct 9988776655",
        "reg": "",
        "customer_no": "CUST-5521",
        "accent": (180, 30, 30),
    },
    {
        "name": "Fjordkraft Komponenter AS",
        "locale": "no",
        "vat": "NO987654321MVA",
        "ccy": "NOK",
        "address": ["Sjøgata 11", "5003 Bergen"],
        "terms": 30,
        "bank": "Kontonr 1234.56.78903",
        "reg": "987 654 321",
        "customer_no": "8802",
        "accent": (20, 60, 140),
    },
]

PRODUCTS = [
    "Steel bracket 40x40",
    "M8 hex bolt, zinc",
    "Bearing 6204-2RS",
    "Hydraulic hose 3/8in",
    "Control panel enclosure",
    "Cable gland M20",
    "Servo motor 750W",
    "Safety relay 24V",
    "Pneumatic cylinder 32x100",
    "Linear guide rail 600mm",
    "Installation labour",
    "Site commissioning",
    "Freight pallet",
    "Terminal block 2.5mm",
    "Proximity sensor M12",
]

LOCALES: dict[str, dict[str, Any]] = {
    "sv": {
        "title": "FAKTURA",
        "number": "Fakturanummer",
        "date": "Fakturadatum",
        "due": "Förfallodatum",
        "delivery": "Leveransdatum",
        "po": "Ert ordernummer",
        "vat": "Momsreg.nr",
        "bill_to": "Fakturamottagare",
        "desc": "Beskrivning",
        "qty": "Antal",
        "unit": "À-pris",
        "amount": "Belopp",
        "net": "Summa netto",
        "discount": "Rabatt",
        "shipping": "Frakt",
        "tax": "Moms",
        "total": "Att betala",
        "rate_col": "Moms %",
        "statement": "KONTOUTDRAG",
        "customer": "Kundnummer",
        "date_short": "Datum",
        "pos": "Pos",
        "sku": "Art.nr",
        "unit_name": "st",
        "hour": "tim",
        "type": "Typ",
        "invoice_word": "Faktura",
        "credit_word": "Kreditnota",
        "balance": "Saldo",
        "status": "Status",
        "paid": "Betald",
        "unpaid": "Obetald",
        "page": "Sida 1 av 1",
        "months": [
            "januari",
            "februari",
            "mars",
            "april",
            "maj",
            "juni",
            "juli",
            "augusti",
            "september",
            "oktober",
            "november",
            "december",
        ],
        "thousands": " ",
        "decimal": ",",
        "ccy_after": True,
    },
    "de": {
        "title": "RECHNUNG",
        "number": "Rechnungsnummer",
        "date": "Rechnungsdatum",
        "due": "Fällig am",
        "delivery": "Lieferdatum",
        "po": "Ihre Bestellnummer",
        "vat": "USt-IdNr.",
        "bill_to": "Rechnungsempfänger",
        "desc": "Bezeichnung",
        "qty": "Menge",
        "unit": "Einzelpreis",
        "amount": "Betrag",
        "net": "Nettobetrag",
        "discount": "Rabatt",
        "shipping": "Versand",
        "tax": "MwSt",
        "total": "Rechnungsbetrag",
        "rate_col": "MwSt %",
        "statement": "KONTOAUSZUG",
        "customer": "Kundennummer",
        "date_short": "Datum",
        "pos": "Pos.",
        "sku": "Art.-Nr.",
        "unit_name": "Stk",
        "hour": "Std",
        "type": "Art",
        "invoice_word": "Rechnung",
        "credit_word": "Gutschrift",
        "balance": "Saldo",
        "status": "Status",
        "paid": "Bezahlt",
        "unpaid": "Offen",
        "page": "Seite 1 von 1",
        "months": [
            "Januar",
            "Februar",
            "März",
            "April",
            "Mai",
            "Juni",
            "Juli",
            "August",
            "September",
            "Oktober",
            "November",
            "Dezember",
        ],
        "thousands": ".",
        "decimal": ",",
        "ccy_after": True,
    },
    "gb": {
        "title": "INVOICE",
        "number": "Invoice number",
        "date": "Invoice date",
        "due": "Due date",
        "delivery": "Delivery date",
        "po": "Your PO",
        "vat": "VAT Reg No",
        "bill_to": "Bill to",
        "desc": "Description",
        "qty": "Qty",
        "unit": "Unit price",
        "amount": "Amount",
        "net": "Net total",
        "discount": "Discount",
        "shipping": "Carriage",
        "tax": "VAT",
        "total": "Total due",
        "rate_col": "VAT %",
        "statement": "STATEMENT OF ACCOUNT",
        "customer": "Account no",
        "date_short": "Date",
        "pos": "#",
        "sku": "Part no",
        "unit_name": "pcs",
        "hour": "hrs",
        "type": "Type",
        "invoice_word": "Invoice",
        "credit_word": "Credit note",
        "balance": "Balance",
        "status": "Status",
        "paid": "Paid",
        "unpaid": "Outstanding",
        "page": "Page 1 of 1",
        "months": [
            "January",
            "February",
            "March",
            "April",
            "May",
            "June",
            "July",
            "August",
            "September",
            "October",
            "November",
            "December",
        ],
        "thousands": ",",
        "decimal": ".",
        "ccy_after": False,
    },
    "us": {
        "title": "INVOICE",
        "number": "Invoice #",
        "date": "Invoice date",
        "due": "Payment due",
        "delivery": "Ship date",
        "po": "PO number",
        "vat": "EIN",
        "bill_to": "Bill to",
        "desc": "Item",
        "qty": "Qty",
        "unit": "Rate",
        "amount": "Amount",
        "net": "Subtotal",
        "discount": "Discount",
        "shipping": "Shipping",
        "tax": "Sales tax",
        "total": "Balance due",
        "rate_col": "Tax %",
        "statement": "ACCOUNT STATEMENT",
        "customer": "Customer ID",
        "date_short": "Date",
        "pos": "#",
        "sku": "SKU",
        "unit_name": "ea",
        "hour": "hrs",
        "type": "Type",
        "invoice_word": "Invoice",
        "credit_word": "Credit memo",
        "balance": "Balance",
        "status": "Status",
        "paid": "Paid",
        "unpaid": "Open",
        "page": "Page 1 of 1",
        "months": [
            "January",
            "February",
            "March",
            "April",
            "May",
            "June",
            "July",
            "August",
            "September",
            "October",
            "November",
            "December",
        ],
        "thousands": ",",
        "decimal": ".",
        "ccy_after": False,
    },
    "no": {
        "title": "FAKTURA",
        "number": "Fakturanr",
        "date": "Fakturadato",
        "due": "Forfallsdato",
        "delivery": "Leveringsdato",
        "po": "Deres ordrenr",
        "vat": "Org.nr",
        "bill_to": "Fakturamottaker",
        "desc": "Beskrivelse",
        "qty": "Antall",
        "unit": "Pris",
        "amount": "Beløp",
        "net": "Sum eks. mva",
        "discount": "Rabatt",
        "shipping": "Frakt",
        "tax": "Mva",
        "total": "Å betale",
        "rate_col": "Mva %",
        "statement": "KONTOUTSKRIFT",
        "customer": "Kundenr",
        "date_short": "Dato",
        "pos": "Pos",
        "sku": "Varenr",
        "unit_name": "stk",
        "hour": "t",
        "type": "Type",
        "invoice_word": "Faktura",
        "credit_word": "Kreditnota",
        "balance": "Saldo",
        "status": "Status",
        "paid": "Betalt",
        "unpaid": "Ubetalt",
        "page": "Side 1 av 1",
        "months": [
            "januar",
            "februar",
            "mars",
            "april",
            "mai",
            "juni",
            "juli",
            "august",
            "september",
            "oktober",
            "november",
            "desember",
        ],
        "thousands": " ",
        "decimal": ",",
        "ccy_after": True,
    },
}

TAX_RATES = {
    "sv": [Decimal(25), Decimal(12)],
    "de": [Decimal(19), Decimal(7)],
    "gb": [Decimal(20), Decimal(5)],
    "us": [Decimal("8.875")],
    "no": [Decimal(25)],
}

# What an operator would tell any pipeline about these documents. The same
# text goes to every approach in the benchmark.
WHO_WE_ARE = (
    f"We are the buyer, {BUYER['name']}, VAT {BUYER['vat']}. Our own name and VAT number "
    "appear on every document and are never the supplier's."
)

INSTRUCTIONS = {
    "invoices": (
        "One row per invoice, not one per line item. subtotal is the net amount after "
        "any discount and before tax, and total is the amount due. Take the invoice "
        f"date, not the delivery date or the due date. {WHO_WE_ARE}"
    ),
    "statement": (
        "One row per invoice listed. Credit notes are not invoices and get no row. "
        f"The supplier is the company that issued the statement. {WHO_WE_ARE}"
    ),
    "invoice_lines": (
        "One row per line item, in the order printed. Skip discount, shipping and total "
        "lines: they are not line items."
    ),
    "expense_claims": "One row for the whole receipt.",
}

# What the schema cannot say and an operator can: the totals must agree, our
# own details can never be the supplier's, a VAT number has a shape, and a
# line's quantity times its price is its total.
RULES = {
    "invoices": [
        {"terms": ["subtotal", "tax"], "equals": "total"},
        {"column": "vat_id", "reject": [BUYER["vat"]]},
        {"column": "supplier", "reject": [BUYER["name"]]},
        {"column": "vat_id", "pattern": "[A-Z]{2}[A-Z0-9]{2,12}"},
    ],
    "invoice_lines": [
        {"factors": ["quantity", "unit_price"], "equals": "line_total"},
    ],
}

# Which visual families a locale's suppliers use. Some are tied to a country's
# conventions: the payment slip is Nordic, the address window is German.
THEMES_BY_LOCALE = {
    "sv": ["classic", "band", "minimal", "slip", "wordmark", "mono", "slip"],
    "de": ["formal_de", "band", "minimal", "wordmark", "mono", "formal_de", "classic"],
    "gb": ["classic", "band", "minimal", "wordmark", "mono", "band"],
    "us": ["us_letter", "wordmark", "mono", "us_letter", "classic"],
    "no": ["slip", "classic", "minimal", "band", "slip"],
}
THEMES_WITHOUT_RATE_COLUMN = {"band", "slip", "mono", "formal_de", "us_letter"}


@dataclass
class Line:
    description: str
    quantity: Decimal
    unit_price: Decimal
    tax_rate: Decimal

    @property
    def total(self) -> Decimal:
        return (self.quantity * self.unit_price).quantize(CENT, ROUND_HALF_UP)

    @property
    def sku(self) -> str:
        digest = hashlib.sha1(self.description.encode()).hexdigest()
        return f"{int(digest[:6], 16) % 90000 + 10000}"

    @property
    def is_service(self) -> bool:
        return "labour" in self.description or "commissioning" in self.description


@dataclass
class Invoice:
    supplier: dict[str, Any]
    number: str
    issued: dt.date
    due: dt.date
    delivered: dt.date
    po: str | None
    lines: list[Line]
    discount_pct: Decimal | None
    shipping: Decimal | None
    printed_total_error: Decimal = Decimal(0)
    date_style: str = "iso"
    theme: str = "classic"

    @property
    def net_lines(self) -> Decimal:
        return sum((line.total for line in self.lines), Decimal(0))

    @property
    def discount(self) -> Decimal:
        if not self.discount_pct:
            return Decimal(0)
        return (self.net_lines * self.discount_pct / 100).quantize(CENT, ROUND_HALF_UP)

    @property
    def subtotal(self) -> Decimal:
        return self.net_lines - self.discount + (self.shipping or Decimal(0))

    @property
    def tax_by_rate(self) -> dict[Decimal, Decimal]:
        """Tax per rate, with the discount spread across lines proportionally."""
        factor = (self.subtotal - (self.shipping or Decimal(0))) / self.net_lines
        by_rate: dict[Decimal, Decimal] = {}
        for line in self.lines:
            by_rate[line.tax_rate] = by_rate.get(line.tax_rate, Decimal(0)) + line.total * factor
        if self.shipping:
            top = max(by_rate)
            by_rate[top] += self.shipping
        return {
            rate: (base * rate / 100).quantize(CENT, ROUND_HALF_UP)
            for rate, base in by_rate.items()
        }

    @property
    def tax(self) -> Decimal:
        return sum(self.tax_by_rate.values(), Decimal(0))

    @property
    def total(self) -> Decimal:
        return self.subtotal + self.tax

    @property
    def printed_total(self) -> Decimal:
        return self.total + self.printed_total_error

    @property
    def locale(self) -> dict[str, Any]:
        return LOCALES[self.supplier["locale"]]

    @property
    def ocr(self) -> str:
        digits = "".join(ch for ch in self.number if ch.isdigit())
        return f"{digits}{int(self.printed_total * 100) % 1000:03d}7"

    def expected_row(self) -> dict[str, Any]:
        return {
            "invoice_number": self.number,
            "supplier": self.supplier["name"],
            "vat_id": self.supplier["vat"],
            "currency": self.supplier["ccy"],
            "subtotal": str(self.subtotal),
            "tax": str(self.tax),
            "total": str(self.printed_total),
            "issued_on": self.issued.isoformat(),
            "due_on": self.due.isoformat(),
            "po_reference": self.po,
        }

    def expected_lines(self) -> list[dict[str, Any]]:
        rows = []
        for index, line in enumerate(self.lines):
            row: dict[str, Any] = {
                "invoice_number": self.number,
                "line_no": index + 1,
                "description": line.description,
                "quantity": str(line.quantity),
                "unit_price": str(line.unit_price),
                "line_total": str(line.total),
            }
            # Only scored where the document prints a rate per line.
            if self.theme not in THEMES_WITHOUT_RATE_COLUMN:
                row["tax_rate"] = str(line.tax_rate)
            rows.append(row)
        return rows


@dataclass
class Case:
    id: str
    file: str
    kind: str
    table: str
    expected: list[dict[str, Any]]
    expected_flags: int = 0
    instructions: str | None = None
    rules: list[dict[str, Any]] = field(default_factory=list)
    notes: str = ""


def money(value: Decimal, loc: dict[str, Any], ccy: str | None = None) -> str:
    sign = "-" if value < 0 else ""
    value = abs(value).quantize(CENT)
    whole, frac = f"{value:f}".split(".")
    groups = []
    while whole:
        groups.insert(0, whole[-3:])
        whole = whole[:-3]
    text = sign + loc["thousands"].join(groups) + loc["decimal"] + frac
    if ccy is None:
        return text
    return f"{text} {ccy}" if loc["ccy_after"] else f"{ccy} {text}"


def number(value: Decimal, loc: dict[str, Any]) -> str:
    text = f"{value.normalize():f}"
    return text.replace(".", loc["decimal"])


def date_text(day: dt.date, loc: dict[str, Any], style: str) -> str:
    if style == "iso":
        return day.isoformat()
    if style == "words":
        return f"{day.day} {loc['months'][day.month - 1]} {day.year}"
    if style == "dmy_dot":
        return day.strftime("%d.%m.%Y")
    if style == "dmy_slash":
        return day.strftime("%d/%m/%Y")
    if style == "mdy":
        return f"{loc['months'][day.month - 1]} {day.day}, {day.year}"
    if style == "mdy_slash":
        return day.strftime("%m/%d/%Y")
    raise ValueError(style)


def make_invoice(rng: random.Random, supplier: dict[str, Any], sequence: int) -> Invoice:
    loc_key = supplier["locale"]
    issued = dt.date(2026, rng.randint(3, 8), rng.randint(1, 28))
    rates = TAX_RATES[loc_key]
    lines = []
    for _ in range(rng.randint(2, 7)):
        description = rng.choice(PRODUCTS)
        if "labour" in description or "commissioning" in description:
            quantity = Decimal(rng.choice(["1.5", "2", "3.5", "8", "12"]))
            unit_price = Decimal(rng.randint(600, 1400))
        else:
            quantity = Decimal(rng.randint(1, 40))
            unit_price = Decimal(rng.randint(1200, 98000)) / 100
        rate = rates[0] if len(rates) == 1 or rng.random() < 0.7 else rates[1]
        lines.append(Line(description, quantity, unit_price, rate))

    prefix = {"sv": "F", "de": "RE-", "gb": "INV-", "us": "", "no": "FA"}[loc_key]
    number_ = f"{prefix}{2026}{sequence:04d}" if loc_key != "us" else f"{10400 + sequence}"
    if loc_key == "gb":
        number_ = f"INV-{2026}-{sequence:04d}"

    date_style = {
        "sv": rng.choice(["iso", "words"]),
        "de": rng.choice(["dmy_dot", "words"]),
        "gb": rng.choice(["words", "dmy_slash"]),
        "us": rng.choice(["mdy", "mdy_slash"]),
        "no": rng.choice(["iso", "dmy_dot"]),
    }[loc_key]

    return Invoice(
        supplier=supplier,
        number=number_,
        issued=issued,
        due=issued + dt.timedelta(days=supplier["terms"]),
        delivered=issued - dt.timedelta(days=rng.randint(1, 9)),
        po=f"PO-2026-{rng.randint(100, 999)}" if rng.random() < 0.6 else None,
        lines=lines,
        discount_pct=Decimal(rng.choice([5, 10, 15])) if rng.random() < 0.3 else None,
        shipping=Decimal(rng.choice([95, 149, 250, 45])) if rng.random() < 0.3 else None,
        date_style=date_style,
        theme=rng.choice(THEMES_BY_LOCALE[loc_key]),
    )


# --- fonts -----------------------------------------------------------------

FONT_DIR = Path("C:/Windows/Fonts")
FONT_FILES = {
    "arial": ("arial.ttf", "arialbd.ttf", "Helvetica"),
    "georgia": ("georgia.ttf", "georgiab.ttf", "Times"),
    "verdana": ("verdana.ttf", "verdanab.ttf", "Helvetica"),
    "calibri": ("calibri.ttf", "calibrib.ttf", "Helvetica"),
    "trebuchet": ("trebuc.ttf", "trebucbd.ttf", "Helvetica"),
    "segoe": ("segoeui.ttf", "segoeuib.ttf", "Helvetica"),
    "tahoma": ("tahoma.ttf", "tahomabd.ttf", "Helvetica"),
    "times": ("times.ttf", "timesbd.ttf", "Times"),
    "consolas": ("consola.ttf", "consolab.ttf", "Courier"),
    "courier": ("cour.ttf", "courbd.ttf", "Courier"),
}


class Pdf(FPDF):
    def __init__(self) -> None:
        super().__init__()
        self.set_auto_page_break(auto=True, margin=15)
        self.set_margins(18, 18, 18)
        self._families: dict[str, str] = {}

    def family(self, key: str) -> str:
        """Register a system font under `key`, or fall back to a core font."""
        if key in self._families:
            return self._families[key]
        regular, bold, core = FONT_FILES[key]
        if (FONT_DIR / regular).exists():
            name = f"Sys{key.capitalize()}"
            self.add_font(name, "", str(FONT_DIR / regular))
            bold_path = FONT_DIR / bold if (FONT_DIR / bold).exists() else FONT_DIR / regular
            self.add_font(name, "B", str(bold_path))
            self._families[key] = name
        else:
            self._families[key] = core
        return self._families[key]

    def use(self, key: str, style: str = "", size: float = 10) -> None:
        self.set_font(self.family(key), style, size)

    def line_text(self, text: str, h: float = 5, align: str = "L") -> None:
        self.cell(0, h, text, align=align, new_x="LMARGIN", new_y="NEXT")

    def ink(self, rgb: tuple[int, int, int] = (0, 0, 0)) -> None:
        self.set_text_color(*rgb)


# --- shared pieces ---------------------------------------------------------


def seller_lines(sup: dict[str, Any], loc: dict[str, Any]) -> list[str]:
    return [sup["name"], *sup["address"], f"{loc['vat']}: {sup.get('ein') or sup['vat']}"]


def buyer_lines(loc: dict[str, Any], with_vat: bool = True) -> list[str]:
    lines = [BUYER["name"], *BUYER["address"]]
    if with_vat:
        lines.append(f"VAT: {BUYER['vat']}")
    return lines


def meta_pairs(inv: Invoice, loc: dict[str, Any], order: str = "default") -> list[tuple[str, str]]:
    ds = inv.date_style
    pairs = {
        "number": (loc["number"], inv.number),
        "date": (loc["date"], date_text(inv.issued, loc, ds)),
        "delivery": (loc["delivery"], date_text(inv.delivered, loc, ds)),
        "due": (loc["due"], date_text(inv.due, loc, ds)),
        "customer": (loc["customer"], inv.supplier["customer_no"]),
        "po": (loc["po"], inv.po or ""),
    }
    orders = {
        "default": ["number", "date", "po", "delivery", "due"],
        "due_first": ["number", "due", "date", "customer", "po", "delivery"],
        "customer_first": ["customer", "number", "date", "delivery", "due", "po"],
    }
    return [pairs[key] for key in orders[order] if key != "po" or inv.po]


def block(
    pdf: Pdf,
    x: float,
    y: float,
    lines: list[str],
    font: str,
    size: float = 10,
    bold_first: bool = False,
    h: float = 5,
    width: float = 0,
) -> float:
    pdf.set_xy(x, y)
    for index, text in enumerate(lines):
        pdf.use(font, "B" if bold_first and index == 0 else "", size)
        pdf.set_x(x)
        pdf.cell(width, h, text, new_x="LMARGIN", new_y="NEXT")
    return pdf.get_y()


def meta_list(
    pdf: Pdf,
    x: float,
    y: float,
    pairs: list[tuple[str, str]],
    font: str,
    label_w: float = 50,
    size: float = 10,
    colon: bool = False,
) -> float:
    pdf.set_xy(x, y)
    for label, value in pairs:
        pdf.set_x(x)
        pdf.use(font, "", size)
        pdf.cell(label_w, 5.5, label + (":" if colon else ""))
        pdf.use(font, "B" if colon else "", size)
        pdf.cell(0, 5.5, value, new_x="LMARGIN", new_y="NEXT")
    return pdf.get_y()


def item_cells(
    inv: Invoice, line: Line, index: int, loc: dict[str, Any], keys: list[str]
) -> list[str]:
    values = {
        "pos": str(index + 1),
        "sku": line.sku,
        "desc": line.description,
        "qty": number(line.quantity, loc),
        "unit_name": loc["hour"] if line.is_service else loc["unit_name"],
        "price": money(line.unit_price, loc),
        "rate": number(line.tax_rate, loc),
        "amount": money(line.total, loc),
    }
    return [values[key] for key in keys]


HEADERS = {
    "pos": "pos",
    "sku": "sku",
    "desc": "desc",
    "qty": "qty",
    "unit_name": "unit_name",
    "price": "unit",
    "rate": "rate_col",
    "amount": "amount",
}
ALIGN = {
    "pos": "R",
    "sku": "L",
    "desc": "L",
    "qty": "R",
    "unit_name": "L",
    "price": "R",
    "rate": "R",
    "amount": "R",
}
UNIT_HEADER = {"sv": "Enhet", "de": "Einheit", "gb": "Unit", "us": "Unit", "no": "Enhet"}


def items_table(
    pdf: Pdf,
    inv: Invoice,
    keys: list[str],
    widths: list[float],
    font: str,
    style: str,
    accent: tuple[int, int, int] = (230, 230, 230),
    size: float = 9.5,
) -> None:
    loc = inv.locale
    headers = [
        UNIT_HEADER[inv.supplier["locale"]] if key == "unit_name" else loc[HEADERS[key]]
        for key in keys
    ]
    pdf.use(font, "B", size)
    if style == "filled":
        pdf.set_fill_color(*accent)
        light = sum(accent) > 400
        pdf.ink((0, 0, 0) if light else (255, 255, 255))
        for width, header, key in zip(widths, headers, keys, strict=True):
            pdf.cell(width, 7, header, align=ALIGN[key], fill=True)
        pdf.ink()
    elif style == "boxed":
        for width, header, key in zip(widths, headers, keys, strict=True):
            pdf.cell(width, 7, header, border=1, align=ALIGN[key])
    elif style == "rule":
        for width, header, key in zip(widths, headers, keys, strict=True):
            pdf.cell(width, 6.5, header, border="B", align=ALIGN[key])
    elif style == "caps":
        pdf.use(font, "", size - 1.5)
        pdf.ink((110, 110, 110))
        for width, header, key in zip(widths, headers, keys, strict=True):
            pdf.cell(width, 6, header.upper(), align=ALIGN[key])
        pdf.ink()
    pdf.ln()
    pdf.use(font, "", size)
    for index, line in enumerate(inv.lines):
        cells = item_cells(inv, line, index, loc, keys)
        if style == "boxed":
            for width, text, key in zip(widths, cells, keys, strict=True):
                pdf.cell(width, 6.5, text, border=1, align=ALIGN[key])
        elif style == "filled" and index % 2 == 1:
            pdf.set_fill_color(243, 243, 243)
            for width, text, key in zip(widths, cells, keys, strict=True):
                pdf.cell(width, 6.5, text, align=ALIGN[key], fill=True)
        else:
            for width, text, key in zip(widths, cells, keys, strict=True):
                pdf.cell(width, 6.5, text, align=ALIGN[key])
        pdf.ln()


def totals_rows(inv: Invoice) -> list[tuple[str, str, bool]]:
    loc = inv.locale
    ccy = inv.supplier["ccy"]
    rows = [(loc["net"], money(inv.net_lines, loc, ccy), False)]
    if inv.discount_pct:
        rows.append(
            (
                f"{loc['discount']} {number(inv.discount_pct, loc)}%",
                money(-inv.discount, loc, ccy),
                False,
            )
        )
    if inv.shipping:
        rows.append((loc["shipping"], money(inv.shipping, loc, ccy), False))
    for rate, amount in sorted(inv.tax_by_rate.items(), reverse=True):
        rows.append((f"{loc['tax']} {number(rate, loc)}%", money(amount, loc, ccy), False))
    rows.append((loc["total"], money(inv.printed_total, loc, ccy), True))
    return rows


def totals(
    pdf: Pdf,
    inv: Invoice,
    x: float,
    font: str,
    label_w: float = 44,
    value_w: float = 38,
    style: str = "plain",
    size: float = 10,
    accent: tuple[int, int, int] = (0, 0, 0),
) -> None:
    for label, value, last in totals_rows(inv):
        pdf.set_x(x)
        if style == "boxed":
            pdf.use(font, "B" if last else "", size)
            pdf.cell(label_w, 6.5, label, border=1)
            pdf.cell(value_w, 6.5, value, border=1, align="R", new_x="LMARGIN", new_y="NEXT")
        elif style == "big_total" and last:
            pdf.ln(2)
            pdf.set_x(x)
            pdf.set_fill_color(*accent)
            pdf.ink((255, 255, 255))
            pdf.use(font, "B", size + 3)
            pdf.cell(label_w, 9, " " + label, fill=True)
            pdf.cell(value_w, 9, value + " ", fill=True, align="R", new_x="LMARGIN", new_y="NEXT")
            pdf.ink()
        else:
            pdf.use(font, "B" if last else "", size + (0.5 if last else 0))
            pdf.cell(label_w, 6, label)
            pdf.cell(value_w, 6, value, align="R", new_x="LMARGIN", new_y="NEXT")


def footer(pdf: Pdf, inv: Invoice, font: str, size: float = 8.5) -> None:
    sup = inv.supplier
    text = {
        "sv": f"Betalning till {sup['bank']}. Ange fakturanummer {inv.number} vid betalning. "
        f"Dröjsmålsränta enligt lag. Org.nr {sup['reg']}. Godkänd för F-skatt.",
        "de": f"Zahlbar innerhalb von {sup['terms']} Tagen ohne Abzug auf {sup['bank']}. "
        f"Verwendungszweck: {inv.number}. {sup['reg']}.",
        "gb": f"Payment within {sup['terms']} days to {sup['bank']}, quoting {inv.number}. "
        f"{sup['reg']}. Registered in England and Wales.",
        "us": f"Net {sup['terms']}. {sup['bank']}. Please reference invoice {inv.number}. "
        f"Thank you for your business.",
        "no": f"Betales til {sup['bank']} innen forfall. Merk betalingen {inv.number}. "
        f"Foretaksregisteret {sup['reg']}.",
    }[sup["locale"]]
    pdf.use(font, "", size)
    pdf.ink((90, 90, 90))
    pdf.multi_cell(0, 4.5, text)
    pdf.ink()


# --- invoice themes --------------------------------------------------------


def theme_classic(pdf: Pdf, inv: Invoice) -> None:
    loc, sup = inv.locale, inv.supplier
    f = "arial"
    pdf.use(f, "B", 18)
    pdf.line_text(loc["title"], 10)
    pdf.ln(2)
    top = pdf.get_y()
    block(pdf, 18, top, seller_lines(sup, loc), f, bold_first=True)
    end = block(pdf, 115, top, [loc["bill_to"], *buyer_lines(loc)], f, bold_first=True)
    pdf.set_y(max(end, top + 36) + 4)
    meta_list(pdf, 18, pdf.get_y(), meta_pairs(inv, loc), f, label_w=55)
    pdf.ln(5)
    items_table(
        pdf, inv, ["desc", "qty", "price", "rate", "amount"], [82, 20, 30, 16, 26], f, "filled"
    )
    pdf.ln(3)
    totals(pdf, inv, 110, f)
    pdf.ln(8)
    footer(pdf, inv, f)


def theme_band(pdf: Pdf, inv: Invoice) -> None:
    loc, sup = inv.locale, inv.supplier
    f = "georgia"
    accent = sup["accent"]
    pdf.set_fill_color(*accent)
    pdf.rect(0, 0, 210, 26, style="F")
    pdf.ink((255, 255, 255))
    pdf.set_xy(18, 7)
    pdf.use(f, "B", 20)
    pdf.cell(100, 12, sup["name"])
    pdf.use(f, "", 14)
    pdf.cell(0, 12, loc["title"], align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.ink()
    pdf.set_y(34)
    top = pdf.get_y()
    block(pdf, 18, top, [loc["bill_to"], *buyer_lines(loc)], f, bold_first=True)
    # Boxed meta grid on the right.
    pdf.set_xy(112, top)
    for label, value in meta_pairs(inv, loc, "due_first"):
        pdf.set_x(112)
        pdf.use(f, "", 8.5)
        pdf.set_fill_color(240, 240, 240)
        pdf.cell(38, 6.5, label, border=1, fill=True)
        pdf.use(f, "B", 9.5)
        pdf.cell(42, 6.5, value, border=1, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)
    pdf.use(f, "", 9)
    pdf.line_text(" · ".join(seller_lines(sup, loc)[1:]), 5)
    pdf.ln(4)
    items_table(
        pdf, inv, ["sku", "desc", "qty", "price", "amount"], [22, 82, 20, 26, 24], f, "boxed"
    )
    pdf.ln(3)
    totals(pdf, inv, 110, f, style="boxed")
    pdf.ln(8)
    footer(pdf, inv, f)
    pdf.use(f, "", 8)
    pdf.set_y(-22)
    pdf.line_text(loc["page"], 5, align="R")


def theme_minimal(pdf: Pdf, inv: Invoice) -> None:
    loc, sup = inv.locale, inv.supplier
    f = "verdana"
    pdf.use(f, "", 8)
    pdf.ink((120, 120, 120))
    pdf.line_text(
        " · ".join([sup["name"], *sup["address"], f"{loc['vat']} {sup.get('ein') or sup['vat']}"]),
        5,
    )
    pdf.ink()
    pdf.ln(10)
    pdf.use(f, "", 22)
    pdf.line_text(f"{loc['title'].capitalize()} {inv.number}", 11)
    pdf.ln(2)
    pdf.use(f, "", 9)
    pairs = [(label, value) for label, value in meta_pairs(inv, loc, "customer_first")]
    pdf.line_text("    ".join(f"{label}: {value}" for label, value in pairs[:3]), 5)
    pdf.line_text("    ".join(f"{label}: {value}" for label, value in pairs[3:]), 5)
    pdf.ln(6)
    pdf.use(f, "", 8)
    pdf.ink((120, 120, 120))
    pdf.line_text(loc["bill_to"].upper(), 4.5)
    pdf.ink()
    block(pdf, 18, pdf.get_y(), buyer_lines(loc), f, size=9.5, h=4.8)
    pdf.ln(8)
    items_table(
        pdf,
        inv,
        ["desc", "qty", "unit_name", "price", "rate", "amount"],
        [74, 16, 16, 28, 16, 24],
        f,
        "caps",
        size=9,
    )
    pdf.set_draw_color(200, 200, 200)
    pdf.line(18, pdf.get_y() + 1, 192, pdf.get_y() + 1)
    pdf.set_draw_color(0, 0, 0)
    pdf.ln(4)
    totals(pdf, inv, 112, f, size=9)
    pdf.ln(10)
    footer(pdf, inv, f, size=7.5)


def theme_slip(pdf: Pdf, inv: Invoice) -> None:
    loc, sup = inv.locale, inv.supplier
    f = "arial"
    pdf.use(f, "B", 26)
    pdf.line_text(loc["title"], 12, align="R")
    top = pdf.get_y() + 2
    block(pdf, 18, top, seller_lines(sup, loc), f, bold_first=True, size=9.5, h=4.8)
    pdf.set_xy(118, top)
    pdf.set_draw_color(160, 160, 160)
    pdf.rect(116, top - 1, 76, 26)
    pdf.set_draw_color(0, 0, 0)
    block(pdf, 119, top + 1, buyer_lines(loc, with_vat=False), f, size=9.5, h=4.8)
    pdf.set_y(top + 32)
    # Horizontal strip of boxes.
    strip = meta_pairs(inv, loc, "customer_first")
    width = 174 / len(strip)
    for label, _ in strip:
        pdf.use(f, "", 7.5)
        pdf.set_fill_color(235, 235, 235)
        pdf.cell(width, 5.5, label, border=1, fill=True)
    pdf.ln()
    for _, value in strip:
        pdf.use(f, "B", 9)
        pdf.cell(width, 7, value, border=1)
    pdf.ln(10)
    items_table(
        pdf,
        inv,
        ["pos", "desc", "qty", "unit_name", "price", "amount"],
        [10, 84, 18, 16, 24, 22],
        f,
        "filled",
        accent=sup["accent"],
    )
    pdf.ln(3)
    totals(pdf, inv, 112, f)
    # Payment slip.
    pdf.set_y(-78)
    pdf.set_draw_color(0, 0, 0)
    pdf.set_dash_pattern(dash=2, gap=2)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.set_dash_pattern()
    pdf.ln(3)
    slip_title = {"sv": "INBETALNING / GIRERING", "no": "GIRO"}.get(sup["locale"], "PAYMENT")
    pdf.use(f, "B", 10)
    pdf.line_text(slip_title, 6)
    pdf.use(f, "", 8.5)
    y = pdf.get_y()
    left = [
        (f"{loc['customer']}: {sup['customer_no']}"),
        (f"{loc['number']}: {inv.number}"),
        (f"{loc['due']}: {date_text(inv.due, loc, 'iso')}"),
        (sup["bank"]),
    ]
    block(pdf, 18, y, left, f, size=8.5, h=4.6)
    pdf.set_xy(118, y)
    pdf.use(f, "", 8.5)
    pdf.cell(
        0, 4.6, "OCR / referens" if sup["locale"] == "sv" else "KID", new_x="LMARGIN", new_y="NEXT"
    )
    pdf.set_x(118)
    pdf.use("consolas", "B", 13)
    pdf.cell(0, 7, inv.ocr, new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(118)
    pdf.use(f, "", 8.5)
    pdf.cell(0, 4.6, loc["total"], new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(118)
    pdf.use("consolas", "B", 13)
    pdf.cell(0, 7, money(inv.printed_total, loc, sup["ccy"]), new_x="LMARGIN", new_y="NEXT")


def theme_wordmark(pdf: Pdf, inv: Invoice) -> None:
    loc, sup = inv.locale, inv.supplier
    f = "trebuchet"
    accent = sup["accent"]
    initials = "".join(word[0] for word in sup["name"].split()[:2]).upper()
    pdf.set_fill_color(*accent)
    pdf.rect(18, 18, 16, 16, style="F")
    pdf.ink((255, 255, 255))
    pdf.set_xy(18, 18)
    pdf.use(f, "B", 13)
    pdf.cell(16, 16, initials, align="C")
    pdf.ink(accent)
    pdf.set_xy(38, 19)
    pdf.use(f, "B", 17)
    pdf.cell(100, 8, sup["name"])
    pdf.ink((110, 110, 110))
    pdf.set_xy(38, 27)
    pdf.use(f, "", 8.5)
    pdf.cell(100, 5, " · ".join([*sup["address"], f"{loc['vat']} {sup.get('ein') or sup['vat']}"]))
    pdf.ink()
    pdf.set_xy(140, 18)
    pdf.use(f, "", 22)
    pdf.ink(accent)
    pdf.cell(52, 10, loc["title"], align="R")
    pdf.ink()
    pdf.set_y(42)
    top = pdf.get_y()
    pdf.use(f, "", 8)
    pdf.ink((110, 110, 110))
    pdf.set_xy(18, top)
    pdf.cell(60, 4.5, loc["bill_to"].upper(), new_x="LMARGIN", new_y="NEXT")
    pdf.ink()
    block(pdf, 18, pdf.get_y(), buyer_lines(loc), f, size=9.5, h=4.8, bold_first=True)
    meta_list(pdf, 118, top, meta_pairs(inv, loc, "due_first"), f, label_w=38, size=9, colon=True)
    pdf.set_y(max(pdf.get_y(), top + 30) + 8)
    items_table(
        pdf,
        inv,
        ["desc", "qty", "price", "rate", "amount"],
        [84, 18, 30, 16, 26],
        f,
        "filled",
        accent=accent,
    )
    pdf.ln(4)
    totals(pdf, inv, 108, f, style="big_total", accent=accent, label_w=46, value_w=38)
    pdf.ln(10)
    footer(pdf, inv, f)


def theme_mono(pdf: Pdf, inv: Invoice) -> None:
    loc, sup = inv.locale, inv.supplier
    f = "courier"
    pdf.use(f, "", 9)

    def row(text: str) -> None:
        pdf.line_text(text, 4.6)

    row(sup["name"].upper().ljust(48) + loc["title"].rjust(26))
    for text in sup["address"]:
        row(text)
    row(f"{loc['vat']}: {sup.get('ein') or sup['vat']}")
    row("=" * 74)
    for label, value in meta_pairs(inv, loc, "customer_first"):
        row(f"{label:.<28}: {value}")
    row("-" * 74)
    row(f"{loc['bill_to'].upper()}:")
    for text in buyer_lines(loc):
        row(f"    {text}")
    row("-" * 74)
    header = (
        f"{loc['pos']:>3} {loc['sku']:<7} {loc['desc']:<28} {loc['qty']:>6} "
        f"{loc['unit']:>12} {loc['amount']:>13}"
    )
    row(header)
    row("-" * 74)
    for index, line in enumerate(inv.lines):
        row(
            f"{index + 1:>3} {line.sku:<7} {line.description[:28]:<28} "
            f"{number(line.quantity, loc):>6} "
            f"{money(line.unit_price, loc):>12} {money(line.total, loc):>13}"
        )
    row("-" * 74)
    for label, value, last in totals_rows(inv):
        text = f"{label:>54} {value:>19}"
        if last:
            pdf.use(f, "B", 9)
        row(text)
        pdf.use(f, "", 9)
    row("=" * 74)
    pdf.ln(3)
    footer(pdf, inv, f, size=8)
    row(loc["page"])


def theme_formal_de(pdf: Pdf, inv: Invoice) -> None:
    loc, sup = inv.locale, inv.supplier
    f = "times"
    pdf.use(f, "B", 15)
    pdf.line_text(sup["name"], 8, align="R")
    pdf.use(f, "", 8)
    pdf.ink((90, 90, 90))
    pdf.line_text(" · ".join([sup["name"], *sup["address"]]), 4.5)
    pdf.ink()
    pdf.ln(8)
    top = pdf.get_y()
    block(pdf, 18, top, buyer_lines(loc, with_vat=False), f, size=10.5, h=5)
    pdf.set_xy(122, top)
    pairs = [
        (loc["date"], date_text(inv.issued, loc, inv.date_style)),
        (loc["customer"], sup["customer_no"]),
        (loc["delivery"], date_text(inv.delivered, loc, inv.date_style)),
        ("Ihre USt-IdNr.", BUYER["vat"]),
        ("Unsere USt-IdNr.", sup["vat"]),
    ]
    if inv.po:
        pairs.insert(2, (loc["po"], inv.po))
    for label, value in pairs:
        pdf.set_x(122)
        pdf.use(f, "", 9)
        pdf.cell(36, 5, label)
        pdf.cell(0, 5, value, new_x="LMARGIN", new_y="NEXT")
    pdf.set_y(max(pdf.get_y(), top + 30) + 10)
    pdf.use(f, "B", 12)
    pdf.line_text(f"Rechnung Nr. {inv.number}", 7)
    pdf.use(f, "", 10)
    pdf.line_text("Sehr geehrte Damen und Herren,", 6)
    pdf.line_text(
        f"für die Lieferung vom {date_text(inv.delivered, loc, 'dmy_dot')} erlauben wir uns, "
        f"Ihnen folgende Positionen in Rechnung zu stellen:",
        6,
    )
    pdf.ln(3)
    items_table(
        pdf,
        inv,
        ["pos", "sku", "desc", "qty", "unit_name", "price", "amount"],
        [10, 18, 70, 16, 14, 24, 22],
        f,
        "rule",
    )
    pdf.ln(3)
    ccy = sup["ccy"]
    pdf.set_x(110)
    rows = [("Zwischensumme", money(inv.net_lines, loc, ccy), False)]
    if inv.discount_pct:
        rows.append(
            (
                f"abzgl. Rabatt {number(inv.discount_pct, loc)}%",
                money(-inv.discount, loc, ccy),
                False,
            )
        )
    if inv.shipping:
        rows.append(("zzgl. Versandkosten", money(inv.shipping, loc, ccy), False))
    rows.append(("Nettobetrag", money(inv.subtotal, loc, ccy), False))
    for rate, amount in sorted(inv.tax_by_rate.items(), reverse=True):
        rows.append((f"zzgl. {number(rate, loc)}% MwSt", money(amount, loc, ccy), False))
    rows.append(("Gesamtbetrag", money(inv.printed_total, loc, ccy), True))
    for label, value, last in rows:
        pdf.set_x(110)
        pdf.use(f, "B" if last else "", 10)
        pdf.cell(46, 6, label)
        pdf.cell(36, 6, value, align="R", border="T" if last else 0, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)
    pdf.use(f, "", 10)
    pdf.line_text(f"Zahlbar bis zum {date_text(inv.due, loc, 'dmy_dot')} ohne Abzug.", 6)
    pdf.line_text("Mit freundlichen Grüßen", 6)
    pdf.line_text(sup["name"], 6)
    pdf.set_y(-30)
    pdf.use(f, "", 7.5)
    pdf.ink((90, 90, 90))
    pdf.multi_cell(
        0,
        4,
        f"{sup['name']} · {' · '.join(sup['address'])} · Geschäftsführer: M. Weber · "
        f"{sup['reg']} · {sup['bank']} · USt-IdNr. {sup['vat']}",
    )
    pdf.ink()


def theme_us_letter(pdf: Pdf, inv: Invoice) -> None:
    loc, sup = inv.locale, inv.supplier
    f = "arial"
    accent = sup["accent"]
    pdf.ink(accent)
    pdf.use(f, "B", 14)
    pdf.cell(110, 7, sup["name"])
    pdf.use(f, "B", 28)
    pdf.cell(0, 12, loc["title"], align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.ink()
    pdf.use(f, "", 9)
    for text in [*sup["address"], f"EIN {sup['ein']}", "billing@greatlakestooling.example"]:
        pdf.line_text(text, 4.6)
    pdf.ln(6)
    top = pdf.get_y()
    pdf.use(f, "B", 9)
    pdf.set_xy(18, top)
    pdf.cell(60, 5, "BILL TO", new_x="LMARGIN", new_y="NEXT")
    block(pdf, 18, pdf.get_y(), buyer_lines(loc), f, size=9.5, h=4.8)
    pdf.use(f, "B", 9)
    pdf.set_xy(80, top)
    pdf.cell(60, 5, "SHIP TO")
    block(pdf, 80, top + 5, BUYER["ship"], f, size=9.5, h=4.8)
    # Meta table right.
    pdf.set_xy(138, top)
    pairs = [
        (loc["number"], inv.number),
        (loc["date"], date_text(inv.issued, loc, inv.date_style)),
        ("Terms", f"Net {sup['terms']}"),
        (loc["due"], date_text(inv.due, loc, inv.date_style)),
        (loc["delivery"], date_text(inv.delivered, loc, inv.date_style)),
        (loc["customer"], sup["customer_no"]),
    ]
    if inv.po:
        pairs.insert(2, (loc["po"], inv.po))
    for label, value in pairs:
        pdf.set_x(138)
        pdf.use(f, "B", 8)
        pdf.set_fill_color(*accent)
        pdf.ink((255, 255, 255))
        pdf.cell(26, 5.5, " " + label, fill=True)
        pdf.ink()
        pdf.use(f, "", 8.5)
        pdf.cell(28, 5.5, value, border=1, align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.set_y(max(pdf.get_y(), top + 32) + 8)
    items_table(
        pdf, inv, ["qty", "desc", "price", "amount"], [18, 100, 28, 28], f, "filled", accent=accent
    )
    pdf.ln(3)
    ccy = sup["ccy"]
    rows = [("Subtotal", money(inv.net_lines, loc, ccy), False)]
    if inv.discount_pct:
        rows.append(
            (f"Discount ({number(inv.discount_pct, loc)}%)", money(-inv.discount, loc, ccy), False)
        )
    if inv.shipping:
        rows.append(("Shipping", money(inv.shipping, loc, ccy), False))
    for rate, amount in inv.tax_by_rate.items():
        rows.append((f"Sales tax ({number(rate, loc)}%)", money(amount, loc, ccy), False))
    rows.append(("Total", money(inv.printed_total, loc, ccy), False))
    rows.append(("Amount paid", money(Decimal(0), loc, ccy), False))
    rows.append(("Balance due", money(inv.printed_total, loc, ccy), True))
    for label, value, last in rows:
        pdf.set_x(118)
        pdf.use(f, "B" if last else "", 10.5 if last else 9.5)
        pdf.cell(40, 6, label)
        pdf.cell(34, 6, value, align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(8)
    pdf.use(f, "", 9)
    pdf.line_text("Thank you for your business.", 5)
    footer(pdf, inv, f, size=8)


THEMES = {
    "classic": theme_classic,
    "band": theme_band,
    "minimal": theme_minimal,
    "slip": theme_slip,
    "wordmark": theme_wordmark,
    "mono": theme_mono,
    "formal_de": theme_formal_de,
    "us_letter": theme_us_letter,
}


def render_invoice(inv: Invoice) -> bytes:
    pdf = Pdf()
    pdf.add_page()
    THEMES[inv.theme](pdf, inv)
    return bytes(pdf.output())


# --- statements ------------------------------------------------------------


@dataclass
class CreditNote:
    number: str
    issued: dt.date
    amount: Decimal
    against: str


def render_statement(
    supplier: dict[str, Any], entries: list[Invoice | CreditNote], period: dt.date, variant: str
) -> bytes:
    """A supplier's statement of account, in one of four shapes.

    plain     one line per invoice with net, tax and total
    typed     a type column, with credit notes as negative lines
    balance   date, reference, amount and running balance only
    status    paid and unpaid marked, all still invoices
    """
    loc = LOCALES[supplier["locale"]]
    ccy = supplier["ccy"]
    pdf = Pdf()
    pdf.add_page()
    f = {"plain": "arial", "typed": "calibri", "balance": "courier", "status": "georgia"}[variant]

    if variant == "balance":
        pdf.use(f, "", 8.5)
        pdf.line_text(supplier["name"].upper().ljust(52) + loc["statement"].rjust(28), 4.6)
        for text in [
            *supplier["address"],
            f"{loc['vat']}: {supplier.get('ein') or supplier['vat']}",
        ]:
            pdf.line_text(text, 4.6)
        pdf.line_text("=" * 80, 4.6)
        pdf.line_text(f"{loc['customer']:.<20}: {supplier['customer_no']}  {BUYER['name']}", 4.6)
        pdf.line_text(f"{'Per':.<20}: {date_text(period, loc, 'iso')}", 4.6)
        pdf.line_text("-" * 80, 4.6)
        pdf.line_text(
            f"{loc['date_short']:<12}{'Ref':<14}{loc['net']:>14}{loc['tax']:>12}{loc['amount']:>14}{loc['balance']:>14}",
            4.6,
        )
        pdf.line_text("-" * 80, 4.6)
        balance = Decimal(0)
        for entry in entries:
            if isinstance(entry, Invoice):
                balance += entry.printed_total
                pdf.line_text(
                    f"{date_text(entry.issued, loc, 'iso'):<12}{entry.number:<14}"
                    f"{money(entry.subtotal, loc):>14}{money(entry.tax, loc):>12}"
                    f"{money(entry.printed_total, loc):>14}{money(balance, loc):>14}",
                    4.6,
                )
            else:
                balance -= entry.amount
                pdf.line_text(
                    f"{date_text(entry.issued, loc, 'iso'):<12}{entry.number:<14}{'':<26}"
                    f"{money(-entry.amount, loc):>14}{money(balance, loc):>14}",
                    4.6,
                )
        pdf.line_text("=" * 80, 4.6)
        pdf.use(f, "B", 8.5)
        pdf.line_text(f"{loc['balance'].upper():>62} {money(balance, loc, ccy):>17}", 4.6)
        return bytes(pdf.output())

    if variant == "status":
        pdf.set_fill_color(*supplier["accent"])
        pdf.rect(0, 0, 210, 22, style="F")
        pdf.ink((255, 255, 255))
        pdf.set_xy(18, 6)
        pdf.use(f, "B", 16)
        pdf.cell(0, 10, loc["statement"])
        pdf.ink()
        pdf.set_y(30)
    else:
        pdf.use(f, "B", 16)
        pdf.line_text(loc["statement"], 9)
    top = pdf.get_y()
    block(
        pdf,
        18,
        top,
        [
            supplier["name"],
            *supplier["address"],
            f"{loc['vat']}: {supplier.get('ein') or supplier['vat']}",
        ],
        f,
        bold_first=True,
    )
    block(
        pdf,
        118,
        top,
        [
            loc["bill_to"],
            BUYER["name"],
            f"VAT: {BUYER['vat']}",
            f"{loc['customer']}: {supplier['customer_no']}",
            f"{date_text(period, loc, 'words')}",
        ],
        f,
        bold_first=True,
    )
    pdf.set_y(max(pdf.get_y(), top + 30) + 6)

    if variant == "plain":
        widths = [30, 26, 26, 30, 28, 34]
        headers = [loc["number"], loc["date"], loc["due"], loc["net"], loc["tax"], loc["total"]]
    elif variant == "typed":
        widths = [24, 30, 26, 30, 28, 36]
        headers = [loc["type"], loc["number"], loc["date"], loc["net"], loc["tax"], loc["total"]]
    else:
        widths = [30, 24, 24, 28, 24, 28, 16]
        headers = [
            loc["number"],
            loc["date"],
            loc["due"],
            loc["net"],
            loc["tax"],
            loc["total"],
            loc["status"],
        ]
    pdf.use(f, "B", 9)
    pdf.set_fill_color(230, 230, 230)
    for width, header in zip(widths, headers, strict=True):
        pdf.cell(width, 6.5, header, border="B", fill=True)
    pdf.ln()
    pdf.use(f, "", 9)
    outstanding = Decimal(0)
    for index, entry in enumerate(entries):
        if isinstance(entry, Invoice):
            paid = variant == "status" and index % 3 == 1
            if not paid:
                outstanding += entry.printed_total
            if variant == "plain":
                cells = [
                    entry.number,
                    date_text(entry.issued, loc, "iso"),
                    date_text(entry.due, loc, "iso"),
                    money(entry.subtotal, loc),
                    money(entry.tax, loc),
                    money(entry.printed_total, loc),
                ]
                aligns = ["L", "L", "L", "R", "R", "R"]
            elif variant == "typed":
                cells = [
                    loc["invoice_word"],
                    entry.number,
                    date_text(entry.issued, loc, "iso"),
                    money(entry.subtotal, loc),
                    money(entry.tax, loc),
                    money(entry.printed_total, loc),
                ]
                aligns = ["L", "L", "L", "R", "R", "R"]
            else:
                cells = [
                    entry.number,
                    date_text(entry.issued, loc, "iso"),
                    date_text(entry.due, loc, "iso"),
                    money(entry.subtotal, loc),
                    money(entry.tax, loc),
                    money(entry.printed_total, loc),
                    loc["paid"] if paid else loc["unpaid"],
                ]
                aligns = ["L", "L", "L", "R", "R", "R", "L"]
        else:
            outstanding -= entry.amount
            cells = [
                loc["credit_word"],
                entry.number,
                date_text(entry.issued, loc, "iso"),
                f"({entry.against})",
                "",
                money(-entry.amount, loc),
            ]
            aligns = ["L", "L", "L", "R", "R", "R"]
        for width, text, align in zip(widths, cells, aligns, strict=True):
            pdf.cell(width, 6, text, align=align)
        pdf.ln()
    pdf.ln(3)
    pdf.use(f, "B", 10)
    pdf.line_text(f"{loc['balance']}: {money(outstanding, loc, ccy)}", 6)
    pdf.use(f, "", 8.5)
    pdf.multi_cell(
        0,
        4.5,
        f"All amounts in {ccy}. Please quote the invoice number with each payment. "
        f"{supplier['bank']}.",
    )
    return bytes(pdf.output())


# --- receipts --------------------------------------------------------------

RECEIPT_MERCHANTS = [
    ("Espresso House Centralen", "meals", ["Cappuccino", "Kanelbulle", "Latte"], 12),
    ("Pressbyrån Odenplan", "meals", ["Smörgås", "Mineralvatten", "Kaffe"], 12),
    ("SJ AB", "travel", ["Stockholm C - Göteborg C, 2 kl", "Platsbokning"], 6),
    ("Taxi Stockholm", "travel", ["Resa Arlanda - Kungsholmen"], 6),
    ("Scandic Continental", "lodging", ["Rum 1 natt", "Frukost"], 12),
    (
        "Clas Ohlson Gallerian",
        "office",
        ["USB-C kabel 2m", "Anteckningsblock A4", "Batterier AA 8-p"],
        25,
    ),
    ("Kjell & Company", "office", ["HDMI-adapter", "Trådlös mus"], 25),
    ("JetBrains s.r.o.", "software", ["PyCharm Professional, 1 month"], 25),
    ("Circle K Norrtull", "travel", ["Diesel 42,17 l", "Spolarvätska"], 25),
    ("Max Burgers Sergels torg", "meals", ["Grand Deluxe meny", "Milkshake"], 12),
    ("Apotek Hjärtat", "other", ["Plåster", "Alvedon 500mg"], 25),
    ("Systembolaget Vasagatan", "other", ["Present till kund"], 25),
]


@dataclass
class Receipt:
    merchant: str
    category: str
    when: dt.datetime
    items: list[tuple[str, Decimal]]
    tax_rate: Decimal
    payment: str
    receipt_number: str

    @property
    def amount(self) -> Decimal:
        return sum((price for _, price in self.items), Decimal(0))

    @property
    def tax(self) -> Decimal:
        return (self.amount * self.tax_rate / (100 + self.tax_rate)).quantize(CENT, ROUND_HALF_UP)

    def expected_row(self) -> dict[str, Any]:
        return {
            "merchant": self.merchant,
            "spent_on": self.when.date().isoformat(),
            "category": self.category,
            "amount": str(self.amount),
            "tax": str(self.tax),
            "currency": "SEK",
            "payment": self.payment,
            "receipt_number": self.receipt_number,
        }


def make_receipt(rng: random.Random, index: int) -> Receipt:
    merchant, category, names, rate = RECEIPT_MERCHANTS[index % len(RECEIPT_MERCHANTS)]
    if category == "meals":
        items = [(name, Decimal(rng.randint(2900, 9900)) / 100) for name in names]
    elif category == "lodging":
        items = [(names[0], Decimal(rng.randint(129000, 219000)) / 100), (names[1], Decimal(0))]
    elif category == "travel":
        items = [(name, Decimal(rng.randint(19900, 89900)) / 100) for name in names]
    else:
        items = [(name, Decimal(rng.randint(4900, 39900)) / 100) for name in names]
    when = dt.datetime(
        2026, rng.randint(4, 8), rng.randint(1, 28), rng.randint(7, 21), rng.randint(0, 59)
    )
    return Receipt(
        merchant=merchant,
        category=category,
        when=when,
        items=items,
        tax_rate=Decimal(rate),
        payment=rng.choice(["card", "card", "card", "cash"]),
        receipt_number=str(rng.randint(10000, 99999)),
    )


def _font(name: str, size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    for candidate in (name, "consola.ttf", "cour.ttf", "DejaVuSansMono.ttf"):
        try:
            return ImageFont.truetype(candidate, size)
        except OSError:
            continue
    return ImageFont.load_default(size)


def render_receipt(receipt: Receipt, rng: random.Random, style: int) -> bytes:
    """A till receipt as a phone photograph, in one of three styles.

    0  monospace thermal print, Swedish
    1  proportional font, English labels, wider paper
    2  narrow paper with quantities, a boxed header, and the VAT in a table
    """
    loc = LOCALES["sv"]
    if style == 0:
        width, font_name, size, paper = 560, "consola.ttf", 22, (248, 246, 240)
    elif style == 1:
        width, font_name, size, paper = 640, "arial.ttf", 22, (252, 252, 250)
    else:
        width, font_name, size, paper = 480, "cour.ttf", 20, (245, 242, 232)
    height = 1000
    image = Image.new("RGB", (width, height), paper)
    draw = ImageDraw.Draw(image)
    font = _font(font_name, size)
    bold = _font(
        font_name.replace(".ttf", "bd.ttf") if "arial" in font_name else font_name, size + 4
    )
    y = 40
    margin = 36

    def line(text: str, f: Any = font, centre: bool = False, right: str | None = None) -> None:
        nonlocal y
        x = margin
        if centre:
            box = draw.textbbox((0, 0), text, font=f)
            x = (width - (box[2] - box[0])) // 2
        draw.text((x, y), text, fill=(30, 30, 30), font=f)
        if right is not None:
            box = draw.textbbox((0, 0), right, font=f)
            draw.text((width - margin - (box[2] - box[0]), y), right, fill=(30, 30, 30), font=f)
        y += int(size * 1.55)

    if style == 2:
        draw.rectangle(
            (margin - 10, y - 10, width - margin + 10, y + 70), outline=(30, 30, 30), width=3
        )
        line(receipt.merchant, bold, centre=True)
        line("Org.nr 556677-8899", centre=True)
        y += 20
    else:
        line(receipt.merchant, bold, centre=True)
        line("Org.nr 556677-8899", centre=True)
        line("Tel 08-123 456 78" if style == 0 else "www.example.se", centre=True)
    y += 10
    if style == 1:
        line(f"Date {receipt.when.strftime('%d/%m/%Y')}", right=receipt.when.strftime("%H:%M"))
        line(f"Receipt no {receipt.receipt_number}", right=f"Till {rng.randint(1, 6)}")
    else:
        line(f"Datum: {receipt.when.strftime('%Y-%m-%d %H:%M')}")
        line(f"Kvittonr: {receipt.receipt_number}")
        line(f"Kassa: {rng.randint(1, 6)}   Kassör: {rng.randint(10, 99)}")
    line("-" * (30 if style != 1 else 38))
    for name, price in receipt.items:
        if style == 2:
            line(f"1 x {name[:18]}", right=money(price, loc))
        else:
            line(name[:24], right=money(price, loc))
    line("-" * (30 if style != 1 else 38))
    total_label = "TOTAL SEK" if style == 1 else "TOTALT SEK"
    line(total_label, bold, right=money(receipt.amount, loc))
    if style == 2:
        line(f"{'Moms%':<8}{'Netto':>10}{'Moms':>10}")
        line(
            f"{number(receipt.tax_rate, loc):<8}{money(receipt.amount - receipt.tax, loc):>10}"
            f"{money(receipt.tax, loc):>10}"
        )
    elif style == 1:
        line(f"VAT {number(receipt.tax_rate, loc)}% incl.", right=money(receipt.tax, loc))
    else:
        line(f"Varav moms {number(receipt.tax_rate, loc)}%", right=money(receipt.tax, loc))
        line("Netto", right=money(receipt.amount - receipt.tax, loc))
    y += 10
    if receipt.payment == "card":
        line(
            ("Paid by card ****" if style == 1 else "Betalt: Kort ****")
            + str(rng.randint(1000, 9999))
        )
        line("Visa Debit  Contactless" if style == 1 else "Visa Debit  Kontaktlös")
        line(f"Ref: {rng.randint(100000, 999999)}")
    else:
        line("Paid in cash" if style == 1 else "Betalt: Kontant")
    y += 10
    line("Thank you!" if style == 1 else "Tack för ditt köp!", centre=True)

    angle = rng.uniform(-3, 3)
    image = image.rotate(
        angle,
        resample=Image.Resampling.BICUBIC,
        expand=True,
        fillcolor=(120, 118, 112) if style != 2 else (60, 70, 80),
    )
    image = image.filter(ImageFilter.GaussianBlur(rng.uniform(0.4, 0.9)))
    buffer = io.BytesIO()
    image.save(buffer, format="JPEG", quality=rng.randint(70, 86))
    return buffer.getvalue()


# --- tabular ---------------------------------------------------------------

TABULAR_HEADERS = {
    "sv": [
        "Fakturanr",
        "Leverantör",
        "Momsreg.nr",
        "Valuta",
        "Netto",
        "Moms",
        "Att betala",
        "Fakturadatum",
        "Förfallodatum",
        "Ert ordernr",
    ],
    "de": [
        "Rechnungsnr.",
        "Lieferant",
        "USt-IdNr.",
        "Währung",
        "Nettobetrag",
        "MwSt",
        "Bruttobetrag",
        "Rechnungsdatum",
        "Fällig am",
        "Bestellnummer",
    ],
    "en": [
        "Invoice Number",
        "Supplier",
        "VAT ID",
        "Currency",
        "Subtotal",
        "Tax",
        "Total",
        "Issued On",
        "Due On",
        "PO Reference",
    ],
}


def tabular_rows(rng: random.Random, count: int, start: int) -> list[Invoice]:
    return [make_invoice(rng, rng.choice(SUPPLIERS), start + index) for index in range(count)]


def write_tabular(
    path: Path, invoices: list[Invoice], headers_key: str, fmt: str, decimal_comma: bool
) -> None:
    headers = TABULAR_HEADERS[headers_key]
    loc = {"thousands": "", "decimal": "," if decimal_comma else "."}
    grid = []
    for invoice in invoices:
        row = invoice.expected_row()
        grid.append(
            [
                row["invoice_number"],
                row["supplier"],
                row["vat_id"] or "",
                row["currency"],
                money(invoice.subtotal, loc),
                money(invoice.tax, loc),
                money(invoice.printed_total, loc),
                row["issued_on"],
                row["due_on"],
                row["po_reference"] or "",
            ]
        )
    if fmt == "xlsx":
        book = Workbook()
        sheet = book.active
        sheet.append(headers)
        for row in grid:
            sheet.append(row)
        book.save(path)
        return
    delimiter = ";" if decimal_comma else ","
    with path.open("w", newline="", encoding="utf-8-sig" if decimal_comma else "utf-8") as handle:
        writer = csv.writer(handle, delimiter=delimiter)
        writer.writerow(headers)
        writer.writerows(grid)


# --- main ------------------------------------------------------------------


def main() -> None:
    rng = random.Random(SEED)
    DATA.mkdir(exist_ok=True)
    for old in DATA.iterdir():
        old.unlink()
    cases: list[Case] = []

    # Single invoices: 45, cycling suppliers. Every ninth has a printed total
    # that does not add up, which the gate is expected to catch.
    sequence = 100
    for index in range(45):
        supplier = SUPPLIERS[index % len(SUPPLIERS)]
        sequence += rng.randint(1, 9)
        invoice = make_invoice(rng, supplier, sequence)
        broken = index % 9 == 4
        if broken:
            invoice.printed_total_error = Decimal(rng.choice([10, 90, 100, -9, 27]))
        name = f"invoice_{index:02d}_{supplier['locale']}_{invoice.theme}.pdf"
        (DATA / name).write_bytes(render_invoice(invoice))
        cases.append(
            Case(
                id=f"inv{index:02d}",
                file=name,
                kind="invoice",
                table="invoices",
                expected=[invoice.expected_row()],
                expected_flags=1 if broken else 0,
                instructions=INSTRUCTIONS["invoices"],
                rules=RULES["invoices"],
                notes=f"theme {invoice.theme}, dates {invoice.date_style}"
                + (", printed total is wrong" if broken else ""),
            )
        )
        if index < 15:
            cases.append(
                Case(
                    id=f"lines{index:02d}",
                    file=name,
                    kind="invoice_lines",
                    table="invoice_lines",
                    expected=invoice.expected_lines(),
                    instructions=INSTRUCTIONS["invoice_lines"],
                    rules=RULES["invoice_lines"],
                    notes=f"theme {invoice.theme}",
                )
            )

    # Statements: 10, three to six invoices each, some with credit notes.
    variants = ["plain", "typed", "balance", "status"]
    for index in range(10):
        supplier = SUPPLIERS[(index * 2) % len(SUPPLIERS)]
        variant = variants[index % len(variants)]
        entries: list[Invoice | CreditNote] = []
        invoices = []
        for _ in range(rng.randint(3, 6)):
            sequence += rng.randint(1, 9)
            invoice = make_invoice(rng, supplier, sequence)
            invoices.append(invoice)
            entries.append(invoice)
        if variant in {"typed", "balance"}:
            target = rng.choice(invoices)
            credit = CreditNote(
                number=f"CN{sequence + 3}",
                issued=target.issued + dt.timedelta(days=5),
                amount=(target.printed_total / 4).quantize(CENT),
                against=target.number,
            )
            entries.insert(entries.index(target) + 1, credit)
        name = f"statement_{index:02d}_{supplier['locale']}_{variant}.pdf"
        (DATA / name).write_bytes(render_statement(supplier, entries, dt.date(2026, 9, 1), variant))
        expected = []
        for invoice in invoices:
            row = {**invoice.expected_row(), "po_reference": None}
            if variant in {"balance", "typed"}:
                # These shapes print no due date.
                row["due_on"] = None
            expected.append(row)
        cases.append(
            Case(
                id=f"stmt{index:02d}",
                file=name,
                kind="statement",
                table="invoices",
                expected=expected,
                instructions=INSTRUCTIONS["statement"],
                rules=RULES["invoices"],
                notes=f"{variant}, {len(invoices)} invoices"
                + (", one credit note" if len(entries) > len(invoices) else ""),
            )
        )

    # Receipts: 12 photographs in three styles.
    for index in range(12):
        receipt = make_receipt(rng, index)
        name = f"receipt_{index:02d}.jpg"
        (DATA / name).write_bytes(render_receipt(receipt, rng, index % 3))
        cases.append(
            Case(
                id=f"rcpt{index:02d}",
                file=name,
                kind="receipt",
                table="expense_claims",
                expected=[receipt.expected_row()],
                instructions=INSTRUCTIONS["expense_claims"],
                notes=f"style {index % 3}",
            )
        )

    # Tabular exports. Headings in three languages, of which only the English
    # set matches column names by spelling.
    specs = [
        ("sv", "csv", True, 50),
        ("de", "csv", True, 200),
        ("en", "xlsx", False, 200),
        ("sv", "xlsx", True, 500),
        ("de", "csv", True, 1000),
        ("sv", "csv", True, 2000),
    ]
    start = 5000
    for index, (lang, fmt, comma, count) in enumerate(specs):
        invoices = tabular_rows(rng, count, start)
        start += count
        name = f"export_{index:02d}_{lang}_{count}.{fmt}"
        write_tabular(DATA / name, invoices, lang, fmt, comma)
        cases.append(
            Case(
                id=f"tab{index:02d}",
                file=name,
                kind="tabular",
                table="invoices",
                expected=[inv.expected_row() for inv in invoices],
                rules=RULES["invoices"],
                notes=f"{lang} headings, {count} rows, {fmt}",
            )
        )

    with (DATA / "manifest.jsonl").open("w", encoding="utf-8") as handle:
        for case in cases:
            handle.write(json.dumps(asdict(case), ensure_ascii=False) + "\n")
    print(f"{len(cases)} cases written to {DATA}")


if __name__ == "__main__":
    main()
